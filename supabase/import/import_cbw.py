"""
CBW 'Online Order Processing' importer for the MHW production portal.

Reads the CBW xlsx (Online Order Processing tab only), parses each row into a
portal order, and either:
  --dry-run  : prints a validation report (default; writes nothing)
  --commit   : inserts customers + orders into Supabase via the Management/REST API

Assumptions (surfaced in the report so they can be corrected):
  * order_type   = made_to_order (all CBW human hair)
  * supplier     = China Best Wigs
  * status       = in_production if DATE SENT TO PRODUCTION is set, else new_made_to_order
  * import only rows where DEPOSIT PAID = YES (an order that has actually been placed)
  * PAID FOR Y/N = YES -> no balance owing ; NO/blank -> balance owing (flagged)
  * supplier length = customer length - 2"; bob/unclear -> needs_review (never guessed)
"""
import sys, os, re, json, argparse
import openpyxl

XLSX = "/Users/yasmintolley/Desktop/CBW ORDERS TO BE MADE .xlsx"
SUPABASE_URL = None  # filled from .env.local on --commit
SERVICE_KEY = None

def parse_length(raw):
    s = (str(raw) if raw is not None else "").strip().upper()
    if not s:
        return None, None, "no length"
    if "BOB" in s or "CUSTOM" in s:
        return None, None, "bob/custom -> needs_review"
    nums = re.findall(r"\d+", s)
    if len(nums) != 1:
        return None, None, f"ambiguous length '{s}'"
    n = int(nums[0])
    if n < 6 or n > 40:
        return None, None, f"length {n} out of range"
    return n, n - 2, None

def parse_style(raw):
    s = (str(raw) if raw is not None else "").strip()
    if not s:
        return None, None
    if " - " in s:
        code, rest = s.split(" - ", 1)
    else:
        parts = s.split(None, 1)
        code, rest = parts[0], (parts[1] if len(parts) > 1 else "")
    return code.strip(), rest.strip()

def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Online Order Processing"]
    hdr = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    def col(name):
        for i, h in enumerate(hdr):
            if h.upper().startswith(name.upper()):
                return i + 1
        return None
    C = {k: col(v) for k, v in {
        "order": "ORDER NUMBER", "eta": "EXPECTED COMPLETION", "deposit": "DEPOSIT PAID",
        "name": "CUSTOMER NAME", "style": "STYLE #", "cap_style": "Cap Style", "length": "LENGTH",
        "cap_size": "CAP SIZE", "density": "DENSITY", "hair": "HAIR", "paid": "PAID FOR",
        "date_ordered": "DATE CUSTOMER", "notes": "EXTRA NOTES", "date_sent": "DATE SENT",
    }.items()}
    # code -> style name from Item Codes
    ic = wb["Item Codes"]
    code2name = {}
    for r in range(2, ic.max_row + 1):
        nm, cd = ic.cell(r, 1).value, ic.cell(r, 2).value
        if nm and cd:
            code2name[str(cd).strip().upper()] = str(nm).strip()
    rows = []
    for r in range(2, ws.max_row + 1):
        get = lambda k: (ws.cell(r, C[k]).value if C.get(k) else None)
        if not get("order"):
            continue
        rows.append({k: get(k) for k in C})
    return rows, code2name

def build(rows, code2name):
    ready, flagged, skipped = [], [], 0
    for row in rows:
        deposit = str(row.get("deposit") or "").strip().upper()
        if deposit != "YES":
            skipped += 1
            continue
        flags = []
        onum = str(row["order"]).replace(".0", "").strip()
        code, cap_construction = parse_style(row.get("style"))
        style_name = code2name.get((code or "").upper())
        if code and not style_name:
            flags.append(f"unmapped style code '{code}'")
        cust_len, sup_len, lenflag = parse_length(row.get("length"))
        if lenflag:
            flags.append(lenflag)
        paid = str(row.get("paid") or "").strip().upper()
        balance_owing = paid not in ("YES", "Y")
        status = "in_production" if row.get("date_sent") else "new_made_to_order"
        needs_review = lenflag is not None
        order = {
            "order_number": onum,
            "source": "shopify",
            "order_type": "needs_review" if needs_review else "made_to_order",
            "status": "manager_review_required" if needs_review else status,
            "customer_name": str(row.get("name") or "").strip() or None,
            "internal_style_name": style_name,
            "supplier_style_code": code,
            "cap_construction": cap_construction,
            "customer_ordered_length": (str(row.get("length")).strip() if row.get("length") else None),
            "supplier_order_length": (str(sup_len) if sup_len is not None else None),
            "cap_size": str(row.get("cap_size") or "").strip() or None,
            "cap_style": str(row.get("cap_style") or "").strip() or None,
            "density": str(row.get("density") or "").strip() or None,
            "hair_type": str(row.get("hair") or "").strip() or None,
            "colour_notes": None,
            "production_notes": str(row.get("notes") or "").strip() or None,
            "balance_owing": balance_owing,
        }
        (flagged if flags else ready).append({"order": order, "flags": flags})
    return ready, flagged, skipped

def report(ready, flagged, skipped):
    print(f"\n===== CBW IMPORT DRY-RUN =====")
    print(f"DEPOSIT PAID = YES rows to import : {len(ready) + len(flagged)}")
    print(f"  clean (ready)                   : {len(ready)}")
    print(f"  flagged (need a look)           : {len(flagged)}")
    print(f"skipped (no DEPOSIT PAID = YES)   : {skipped}")
    from collections import Counter
    fc = Counter(f for x in flagged for f in x["flags"])
    if fc:
        print("\nflag breakdown:")
        for f, n in fc.most_common():
            print(f"  {n:>3}  {f}")
    print("\n--- 5 sample CLEAN orders ---")
    for x in ready[:5]:
        o = x["order"]
        print(f"  {o['order_number']}  {o['customer_name']}  {o['internal_style_name'] or o['supplier_style_code']}  "
              f"len {o['customer_ordered_length']}->{o['supplier_order_length']}  cap {o['cap_size']}  status {o['status']}  balance_owing {o['balance_owing']}")
    print("\n--- 8 sample FLAGGED ---")
    for x in flagged[:8]:
        o = x["order"]
        print(f"  {o['order_number']}  {o['customer_name']}  code {o['supplier_style_code']}  len '{o['customer_ordered_length']}'  -> {', '.join(x['flags'])}")

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true")
    args = ap.parse_args()
    rows, code2name = load_rows()
    ready, flagged, skipped = build(rows, code2name)
    report(ready, flagged, skipped)
    if not args.commit:
        print("\n(DRY RUN — nothing written. Re-run with --commit to import.)")

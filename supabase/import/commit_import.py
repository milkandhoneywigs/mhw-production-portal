"""Import the reviewed order list into the live portal DB.

Source of truth = the (human-edited) review file on the Desktop.
Default = validate only. Pass --commit to write to Supabase.
Idempotent: an order_number that already exists is skipped, so re-running is safe.
"""
import sys, os, re, json, argparse
import openpyxl, requests

REVIEW = "/Users/yasmintolley/Desktop/MHW PORTAL IMPORT - REVIEW.xlsx"
ENV = "/Users/yasmintolley/mhw-production-portal/.env.local"

VALID_STATUS = {
 'new_ready_made_order','supplier_notified','awaiting_dhl_tracking','tracking_uploaded','customer_notified',
 'new_made_to_order','awaiting_supplier_confirmation','invoice_uploaded','payment_required','payment_paid',
 'in_production','production_update_due','production_complete','balance_payment_required','balance_paid',
 'shipped_to_showroom','arrived_at_showroom','qc_required','qc_passed','ready_to_dispatch',
 'dispatched_to_customer','delayed_at_risk','manager_review_required','completed'}
VALID_TYPE = {'ready_made','made_to_order','stock','needs_review'}

def env(k):
    for line in open(ENV):
        if line.startswith(k+'='): return line.split('=',1)[1].strip()
    return None

def load_rows():
    wb = openpyxl.load_workbook(REVIEW, data_only=True)
    ws = wb.active
    hdr = [str(ws.cell(1,c).value or '').strip() for c in range(1, ws.max_column+1)]
    rows = []
    for r in range(2, ws.max_row+1):
        vals = {hdr[c-1]: ws.cell(r,c).value for c in range(1, ws.max_column+1)}
        if not (vals.get('order_number') and str(vals['order_number']).strip()): continue
        rows.append(vals)
    return rows

def val(v):
    s = ('' if v is None else str(v)).strip()
    return s or None

def validate(rows):
    issues = []
    seen = set()
    for i, r in enumerate(rows, 2):
        onum = val(r.get('order_number'))
        if onum in seen: issues.append(f"row {i}: duplicate order_number {onum}")
        seen.add(onum)
        if not val(r.get('customer_name')): issues.append(f"row {i} ({onum}): missing customer_name")
        st = val(r.get('status'))
        if st and st not in VALID_STATUS: issues.append(f"row {i} ({onum}): invalid status '{st}'")
        ot = val(r.get('order_type'))
        if ot and ot not in VALID_TYPE: issues.append(f"row {i} ({onum}): invalid order_type '{ot}'")
    return issues

def main():
    ap = argparse.ArgumentParser(); ap.add_argument('--commit', action='store_true'); a = ap.parse_args()
    rows = load_rows()
    issues = validate(rows)
    print(f"rows to import: {len(rows)} | validation issues: {len(issues)}")
    for x in issues[:30]: print("  !", x)
    if not a.commit:
        print("\n(VALIDATE ONLY — re-run with --commit to write.)"); return
    if issues:
        print("\nRefusing to commit while there are validation issues above. Fix them first."); return

    URL, KEY = env('NEXT_PUBLIC_SUPABASE_URL'), env('SUPABASE_SERVICE_ROLE_KEY')
    H = {"apikey": KEY, "Authorization": f"Bearer {KEY}", "Content-Type": "application/json"}
    # ensure supplier
    sup = requests.get(f"{URL}/rest/v1/suppliers?name=eq.China Best Wigs&select=id", headers=H).json()
    if sup: supplier_id = sup[0]['id']
    else:
        supplier_id = requests.post(f"{URL}/rest/v1/suppliers", headers={**H,"Prefer":"return=representation"},
            data=json.dumps({"name":"China Best Wigs","active":True})).json()[0]['id']
    print("supplier China Best Wigs:", supplier_id)

    created = skipped = failed = 0
    for r in rows:
        onum = val(r.get('order_number'))
        exists = requests.get(f"{URL}/rest/v1/orders?order_number=eq.{onum}&select=id", headers=H).json()
        if exists: skipped += 1; continue
        # customer
        cust = requests.post(f"{URL}/rest/v1/customers", headers={**H,"Prefer":"return=representation"},
            data=json.dumps({"full_name": val(r.get('customer_name')), "email": val(r.get('email')),
                "phone": val(r.get('phone')), "shipping_address_line1": val(r.get('shipping_address'))}))
        if cust.status_code >= 300: failed += 1; print("cust fail", onum, cust.text[:120]); continue
        cid = cust.json()[0]['id']
        notes = []
        if str(val(r.get('fully_paid')) or '').upper() == 'NO': notes.append("Balance owing (per import).")
        if val(r.get('flags')): notes.append("Import flags: " + val(r.get('flags')))
        order = {
            "order_number": onum, "source": val(r.get('source')) or 'shopify',
            "customer_id": cid, "supplier_id": supplier_id,
            "order_type": val(r.get('order_type')) or 'made_to_order',
            "status": val(r.get('status')) or 'new_made_to_order',
            "customer_facing_product_name": val(r.get('style_name')),
            "internal_style_name": val(r.get('style_name')),
            "supplier_style_code": val(r.get('supplier_style_code')),
            "customer_ordered_length": val(r.get('customer_ordered_length')),
            "supplier_order_length": val(r.get('supplier_order_length')),
            "cap_style": val(r.get('cap_style')), "cap_size": val(r.get('cap_size')),
            "density": val(r.get('density')), "hair_type": val(r.get('hair_type')) or 'human hair',
            "colour_notes": val(r.get('colour_notes')), "production_notes": val(r.get('production_notes')),
            "internal_notes": " ".join(notes) or None, "shipping_destination": "mhw_showroom",
        }
        od = val(r.get('order_date'))
        if od and re.match(r"\d{4}-\d{2}-\d{2}", od): order["date_ordered"] = od[:10]
        res = requests.post(f"{URL}/rest/v1/orders", headers=H, data=json.dumps(order))
        if res.status_code >= 300: failed += 1; print("order fail", onum, res.text[:150])
        else: created += 1

    print(f"\n=== IMPORT COMPLETE ===\ncreated: {created} | skipped (already existed): {skipped} | failed: {failed}")

if __name__ == "__main__": main()

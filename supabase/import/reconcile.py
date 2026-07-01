"""Reconcile Shopify open orders (orders_export_1.csv) vs the CBW production
tracker (Online Order Processing tab). Read-only — reports only."""
import csv, re, sys
from collections import defaultdict
import openpyxl

SHOP = "/Users/yasmintolley/Desktop/orders_export_1.csv"
XLSX = "/Users/yasmintolley/Desktop/CBW ORDERS TO BE MADE .xlsx"

def norm(n):
    if n is None: return None
    s = str(n).strip()
    m = re.match(r"^(\d+)\.0+$", s)   # spreadsheet stores '31312.0' -> 31312
    if m: s = m.group(1)
    s = re.sub(r"[^0-9]", "", s)
    return s or None

# --- Shopify: group rows by order Name ---
orders = {}
with open(SHOP, encoding="utf-8-sig") as fh:
    for row in csv.DictReader(fh):
        name = row.get("Name")
        if not name: continue
        o = orders.setdefault(name, {"name": name, "fulfil": row.get("Fulfillment Status"),
            "fin": row.get("Financial Status"), "cust": row.get("Shipping Name") or row.get("Billing Name"),
            "created": row.get("Created at"), "items": []})
        li = (row.get("Lineitem name") or "").strip()
        if li: o["items"].append((li, (row.get("Lineitem sku") or "").strip()))

def is_open(o):
    f = (o["fulfil"] or "").strip().lower()
    return f in ("", "unfulfilled", "partial", "partially fulfilled")

def is_human_hair(o):
    # a CBW order = at least one line item that is a "human hair wig" (excludes
    # accessories, haircare, tape, stands, and synthetic lace-front/fringe wigs).
    return any("human hair" in (n or "").lower() for n, _ in o["items"])

open_orders = {norm(o["name"]): o for o in orders.values() if is_open(o)}
open_hh = {k: o for k, o in open_orders.items() if is_human_hair(o)}

# --- Sheet: order numbers on the Online Order Processing tab ---
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Online Order Processing"]
sheet_all, sheet_outstanding = set(), set()
for r in range(2, ws.max_row + 1):
    onum = ws.cell(r, 1).value
    if not onum: continue
    n = norm(onum)
    if not n: continue  # INSTORE / non-numeric -> skip matching
    sheet_all.add(n)
    if str(ws.cell(r, 3).value or "").strip().upper() == "YES":  # DEPOSIT PAID
        sheet_outstanding.add(n)

# --- Reconcile ---
print("=== SHOPIFY EXPORT ===")
print(f"  distinct orders: {len(orders)} | open: {len(open_orders)} | open + human-hair: {len(open_hh)}")
print(f"=== SHEET (Online Order Processing) ===")
print(f"  numeric orders: {len(sheet_all)} | outstanding (DEPOSIT PAID=YES): {len(sheet_outstanding)}")

in_both = sorted(set(open_hh) & sheet_all)
shop_not_sheet = sorted(set(open_hh) - sheet_all)          # open HH in Shopify, missing from tracker
sheet_not_open = sorted(sheet_outstanding - set(open_orders))  # outstanding on sheet, not open in Shopify

print(f"\n✅ In BOTH (open HH in Shopify + on tracker): {len(in_both)}")
print(f"⚠️  In SHOPIFY (open human-hair) but NOT on tracker: {len(shop_not_sheet)}")
for k in shop_not_sheet:
    o = open_hh[k]
    items = "; ".join(n for n, _ in o["items"][:2])
    print(f"     {o['name']}  {o['cust']}  [{items[:50]}]  ({o['fin']}, {o['created'][:10] if o['created'] else '?'})")
print(f"\n⚠️  OUTSTANDING on tracker but NOT open in Shopify (fulfilled/cancelled/stale?): {len(sheet_not_open)}")
print("     " + ", ".join(sheet_not_open))
# also: non-human-hair open orders in shopify not on sheet (FYI, likely synthetic/accessories)
open_other = sorted(set(open_orders) - set(open_hh) - sheet_all)
print(f"\nℹ️  Other open Shopify orders (not human-hair, not on tracker): {len(open_other)} (synthetic/accessories — not for CBW import)")

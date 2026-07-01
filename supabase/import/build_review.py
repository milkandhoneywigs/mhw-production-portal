"""Build the REVIEW export for the CBW import.

Final outstanding list = every Shopify OPEN + HUMAN-HAIR order that is NOT a
SAMPLE, enriched with tracker production detail where the order number matches.
  * already-fulfilled (not open in Shopify) -> excluded automatically
  * SAMPLE orders -> excluded (hard rule)
Writes ~/Desktop/MHW PORTAL IMPORT - REVIEW.xlsx for sign-off. Imports nothing.
"""
import csv, re
from collections import defaultdict
import openpyxl

SHOP = "/Users/yasmintolley/Desktop/orders_export_1.csv"
XLSX = "/Users/yasmintolley/Desktop/CBW ORDERS TO BE MADE .xlsx"
OUT  = "/Users/yasmintolley/Desktop/MHW PORTAL IMPORT - REVIEW.xlsx"

def norm(n):
    if n is None: return None
    s = str(n).strip()
    m = re.match(r"^(\d+)\.0+$", s)
    if m: s = m.group(1)
    s = re.sub(r"[^0-9]", "", s)
    return s or None

# ---- Item Codes: name<->code ----
wb = openpyxl.load_workbook(XLSX, data_only=True)
ic = wb["Item Codes"]
name2code = {}
for r in range(2, ic.max_row + 1):
    nm, cd = ic.cell(r, 1).value, ic.cell(r, 2).value
    if nm and cd: name2code[str(nm).strip().upper()] = str(cd).strip()

# ---- Tracker rows by order number ----
ws = wb["Online Order Processing"]
tracker = {}
for r in range(2, ws.max_row + 1):
    onum = ws.cell(r, 1).value
    n = norm(onum)
    if not n: continue
    tracker[n] = {
        "style": ws.cell(r, 5).value, "cap_style": ws.cell(r, 6).value, "length": ws.cell(r, 7).value,
        "cap_size": ws.cell(r, 8).value, "density": ws.cell(r, 9).value, "hair": ws.cell(r, 10).value,
        "paid": ws.cell(r, 11).value, "notes": ws.cell(r, 13).value, "deposit": ws.cell(r, 3).value,
    }

# ---- Shopify: group open human-hair, non-sample orders ----
orders = {}
with open(SHOP, encoding="utf-8-sig") as fh:
    for row in csv.DictReader(fh):
        name = row.get("Name")
        if not name: continue
        o = orders.setdefault(name, {"name": name, "fulfil": row.get("Fulfillment Status"),
            "fin": row.get("Financial Status"), "email": row.get("Email"),
            "cust": row.get("Shipping Name") or row.get("Billing Name"),
            "phone": row.get("Shipping Phone") or row.get("Billing Phone"),
            "addr1": row.get("Shipping Address1"), "addr2": row.get("Shipping Address2"),
            "city": row.get("Shipping City"), "zip": row.get("Shipping Zip"),
            "prov": row.get("Shipping Province"), "created": row.get("Created at"), "items": []})
        li = (row.get("Lineitem name") or "").strip()
        if li: o["items"].append(li)

def is_open(o): return (o["fulfil"] or "").strip().lower() in ("", "unfulfilled", "partial", "partially fulfilled")
def hh_item(o):
    for n in o["items"]:
        if "human hair" in n.lower(): return n
    return None
def is_sample(o): return any("sample" in n.lower() for n in o["items"])

def parse_title(title):
    """Pull style / colour / cap / length from a Shopify wig product title."""
    parts = [p.strip() for p in title.split(" - ")]
    style = parts[0] if parts else ""
    colour = parts[1] if len(parts) > 1 else ""
    lm = re.search(r"(\d+)\s*(?:''|\"|inch|in\b)", title, re.I)
    length = lm.group(1) if lm else ""
    cap = ""
    for k in ["13 x 6", "13x6", "Full Lace", "Medical Cap", "Seamless", "Lace Front", "Petite", "Regular", "Medium", "Large"]:
        if k.lower() in title.lower(): cap = k; break
    return style, colour, length, cap

def sup_len(length_raw):
    s = str(length_raw or "").upper()
    if "BOB" in s or "CUSTOM" in s: return "", "bob/custom -> needs_review"
    nums = re.findall(r"\d+", s)
    if len(nums) != 1: return "", ("no length" if not nums else "ambiguous length")
    n = int(nums[0])
    return (str(n - 2), None) if 6 <= n <= 40 else ("", "length out of range")

rows_out, n_sample, n_matched, n_missing = [], 0, 0, 0
for o in orders.values():
    if not is_open(o): continue
    if not hh_item(o): continue
    if is_sample(o):   n_sample += 1; continue   # HARD RULE: exclude samples
    onum = norm(o["name"])
    t = tracker.get(onum)
    on_tracker = "YES" if t else "NO"
    if t: n_matched += 1
    else: n_missing += 1
    s_style, s_colour, s_length, s_cap = parse_title(hh_item(o))
    # production fields: prefer tracker (what was actually ordered from CBW), else Shopify title
    if t:
        style_raw = str(t["style"] or "")
        code = (style_raw.split(" - ")[0].strip() if " - " in style_raw else style_raw.split()[0]) if style_raw else ""
        length = t["length"] or s_length
        cap_size = t["cap_size"]; cap_style = t["cap_style"] or s_cap
        density = t["density"]; hair = t["hair"]; notes = t["notes"]
        paid = str(t["paid"] or "").strip().upper()
    else:
        code = name2code.get(s_style.upper(), "")
        length = s_length; cap_size = ""; cap_style = s_cap; density = ""; hair = ""; notes = ""; paid = ""
    style_name = next((nm for nm, cd in name2code.items() if cd.upper() == str(code).upper()), "") or (s_style if not t else "")
    slen, lenflag = sup_len(length)
    flags = []
    if code and not style_name: flags.append("style code not in Item Codes")
    if not code: flags.append("no style code")
    if lenflag: flags.append(lenflag)
    status = "in_production" if t else "new_made_to_order"
    order_type = "needs_review" if lenflag else "made_to_order"
    rows_out.append([
        onum, "shopify", on_tracker, order_type, ("manager_review_required" if lenflag else status),
        o["cust"], o["email"], o["phone"],
        ", ".join(x for x in [o["addr1"], o["addr2"], o["city"], o["prov"], o["zip"]] if x),
        (o["created"] or "")[:10], style_name, code, cap_style, s_colour,
        str(length or ""), slen, cap_size, density, hair, notes,
        "YES" if paid in ("YES","Y") else "NO", "; ".join(flags),
    ])

rows_out.sort(key=lambda r: r[0])
HDR = ["order_number","source","on_tracker","order_type","status","customer_name","email","phone",
       "shipping_address","order_date","style_name","supplier_style_code","cap_style","colour_notes",
       "customer_ordered_length","supplier_order_length","cap_size","density","hair_type",
       "production_notes","fully_paid","flags"]
out = openpyxl.Workbook(); sh = out.active; sh.title = "Import Review"
sh.append(HDR)
for r in rows_out: sh.append(r)
out.save(OUT)

print(f"FINAL import list: {len(rows_out)} orders")
print(f"  from tracker (matched):        {n_matched}")
print(f"  new from Shopify (not tracked): {n_missing}")
print(f"  SAMPLE orders EXCLUDED:          {n_sample}")
print(f"  (already-fulfilled excluded automatically — not in Shopify's open export)")
print(f"  flagged rows (need a look):     {sum(1 for r in rows_out if r[-1])}")
print(f"\nWritten to: {OUT}")
